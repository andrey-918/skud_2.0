import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
from collections import defaultdict
import sqlite3
from db.reports import get_all_attendance_records

class ReportsTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.create_ui()

    def create_ui(self):
        # Threshold for percentage coloring
        threshold_frame = ttk.LabelFrame(self, text="Настройки")
        threshold_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(threshold_frame, text="Порог процента для окраски (%):").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.percentage_threshold = ttk.Entry(threshold_frame)
        self.percentage_threshold.insert(0, "50")
        self.percentage_threshold.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)

        # Export button for full report
        ttk.Button(self, text="Экспорт общей таблицы в Excel", command=self.export_all_report_to_excel).pack(pady=20)

    def export_all_report_to_excel(self):
        """Export comprehensive attendance report to Excel in wide format"""
        try:
            # Get all attendance records
            records = get_all_attendance_records()
            if not records:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
                return

            # Group records by student
            student_data = defaultdict(lambda: defaultdict(dict))
            student_info = {}

            for record in records:
                student_name = record['student_name']
                meal_id = record['meal_id']
                day_of_week = record['day_of_week']
                status = record['status']

                # Calculate meal type: 0=Breakfast, 1=Lunch, 2=Dinner
                meal_type = (meal_id - 1) % 3
                student_data[student_name][day_of_week][meal_type] = status

                # Store student info (assuming all records have same info)
                if student_name not in student_info:
                    # Get student details
                    conn = sqlite3.connect('skud.db')
                    cursor = conn.cursor()
                    cursor.execute('SELECT id, card_id, group_name FROM students WHERE name = ?', (student_name,))
                    result = cursor.fetchone()
                    conn.close()
                    if result:
                        student_info[student_name] = {'id': result[0], 'card_id': result[1], 'group': result[2]}

            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Общая таблица посещаемости"

            # Headers similar to report.xlsx
            ws['A1'] = "График обеспечением питания студентов Морского института"
            ws['A5'] = "Дата"
            ws['C5'] = f"{datetime.now().strftime('%d.%m.%Y')}"
            ws['A6'] = "День недели"
            ws['D6'] = "Понедельник"
            ws['G6'] = "Вторник"
            ws['J6'] = "Среда"
            ws['M6'] = "Четверг"
            ws['P6'] = "Пятница"
            ws['S6'] = "Суббота"
            ws['A7'] = "Локация"
            ws['D7'] = "Гоголя"
            ws['G7'] = "Гоголя"
            ws['J7'] = "Гоголя"
            ws['M7'] = "Гоголя"
            ws['P7'] = "Гоголя"
            ws['S7'] = "Гоголя"
            ws['A8'] = "Учебная группа"
            ws['B8'] = "№п/п"
            ws['C8'] = "ФИО"
            # Meal headers: З (Breakfast), О (Lunch), У (Dinner)
            col = 4
            for day in range(7):
                ws.cell(row=8, column=col).value = "З"
                ws.cell(row=8, column=col+1).value = "О"
                ws.cell(row=8, column=col+2).value = "У"
                col += 3

            # Additional columns
            ws.cell(row=8, column=22).value = "Всего заявок"
            ws.cell(row=8, column=23).value = "Питание по заявке"
            ws.cell(row=8, column=24).value = "Питание без заявки"
            ws.cell(row=8, column=25).value = "Процент посещений"

            # Data rows
            row = 9
            total_registered_sum = 0
            came_with_reg_sum = 0
            came_without_reg_sum = 0
            breakfast_sums = [0] * 7  # For each day
            lunch_sums = [0] * 7
            dinner_sums = [0] * 7
            for idx, student_name in enumerate(sorted(student_data.keys()), start=1):
                info = student_info.get(student_name, {'group': '', 'id': '', 'card_id': ''})
                ws.cell(row=row, column=1).value = info['group'] or ""
                ws.cell(row=row, column=2).value = idx
                ws.cell(row=row, column=3).value = student_name

                total_registered = 0
                came_with_registration = 0
                came_without_registration = 0

                col = 4
                for day in range(7):
                    for meal_type in range(3):  # 0: Breakfast, 1: Lunch, 2: Dinner
                        status = student_data[student_name][day].get(meal_type, 'not_registered')
                        cell = ws.cell(row=row, column=col)
                        if status == 'came':
                            cell.value = 1
                            came_with_registration += 1
                            total_registered += 1
                        elif status == 'didnt_come':
                            cell.value = 1
                            cell.fill = yellow_fill
                            total_registered += 1
                        elif status == 'came_without_registration':
                            cell.value = 1
                            came_without_registration += 1
                        elif status == 'not_registered':
                            cell.value = 0
                        # Sum registered for meals
                        if status in ['came', 'didnt_come']:
                            if meal_type == 0:
                                breakfast_sums[day] += 1
                            elif meal_type == 1:
                                lunch_sums[day] += 1
                            elif meal_type == 2:
                                dinner_sums[day] += 1
                        col += 1

                # Calculate total registered (didnt_come + came_with_registration)
                total_registered += came_with_registration

                # Additional columns
                cell_22 = ws.cell(row=row, column=22)
                cell_22.value = total_registered

                cell_23 = ws.cell(row=row, column=23)
                cell_23.value = came_with_registration

                cell_24 = ws.cell(row=row, column=24)
                cell_24.value = came_without_registration

                cell_25 = ws.cell(row=row, column=25)
                if total_registered > 0:
                    percentage = (came_with_registration / total_registered) * 100
                    cell_25.value = round(percentage, 2)
                    try:
                        threshold = float(self.percentage_threshold.get())
                        if percentage < threshold:
                            cell_25.fill = red_fill
                    except ValueError:
                        pass  # Ignore invalid threshold
                else:
                    cell_25.value = 0

                # Sum for totals
                total_registered_sum += total_registered
                came_with_reg_sum += came_with_registration
                came_without_reg_sum += came_without_registration

                row += 1

            # Add meal summary row
            ws.cell(row=row, column=3).value = "Сводка"
            col = 4
            for day in range(6):  # Exclude Sunday
                ws.cell(row=row, column=col).value = breakfast_sums[day]
                ws.cell(row=row, column=col+1).value = lunch_sums[day]
                ws.cell(row=row, column=col+2).value = dinner_sums[day]
                col += 3
            ws.cell(row=row, column=22).value = total_registered_sum
            ws.cell(row=row, column=23).value = came_with_reg_sum
            ws.cell(row=row, column=24).value = came_without_reg_sum

            filename = f"отчеты/Общая_таблица_посещаемости_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            wb.save(filename)

            messagebox.showinfo("Успех", f"Отчет экспортирован в {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать отчет: {str(e)}")
