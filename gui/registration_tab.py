import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import openpyxl
from db.student import finding_card, add_registration, find_student_by_name_group

class RegistrationTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.create_ui()

    def create_ui(self):
        # Student selection
        student_frame = ttk.LabelFrame(self, text="Выбор студента")
        student_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(student_frame, text="ID карты студента:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.reg_card_id = ttk.Entry(student_frame)
        self.reg_card_id.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)

        ttk.Button(student_frame, text="Найти студента", command=self.find_student_for_registration).grid(row=0, column=2, padx=5, pady=5)

        self.student_info_label = ttk.Label(student_frame, text="")
        self.student_info_label.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky=tk.W)

        # Day selection
        day_frame = ttk.LabelFrame(self, text="Выбор дня")
        day_frame.pack(fill=tk.X, padx=10, pady=10)

        self.day_var = tk.StringVar(value="-1")
        days = [("-1", "Все дни"), ("0", "Понедельник"), ("1", "Вторник"), ("2", "Среда"),
                ("3", "Четверг"), ("4", "Пятница"), ("5", "Суббота"), ("6", "Воскресенье")]

        for i, (val, text) in enumerate(days):
            ttk.Radiobutton(day_frame, text=text, variable=self.day_var, value=val).grid(row=i//4, column=i%4, padx=5, pady=5, sticky=tk.W)

        # Meals selection
        meals_frame = ttk.LabelFrame(self, text="Выбор приемов пищи")
        meals_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.meals_listbox = tk.Listbox(meals_frame, selectmode=tk.MULTIPLE, height=10)
        self.meals_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(meals_frame, orient=tk.VERTICAL, command=self.meals_listbox.yview)
        self.meals_listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_frame, text="Загрузить приемы пищи", command=self.load_meals_for_registration).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Зарегистрировать", command=self.register_student).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Импорт из XLSX", command=self.import_students_from_xlsx).pack(side=tk.LEFT, padx=5)

    def find_student_for_registration(self):
        """Find student by card ID for registration"""
        card_id_str = self.reg_card_id.get()
        if not card_id_str:
            messagebox.showwarning("Предупреждение", "Введите ID карты")
            return

        try:
            card_id = card_id_str
            student_id = finding_card(card_id)
            if student_id:
                # Get student name
                conn = sqlite3.connect('skud.db')
                cursor = conn.cursor()
                cursor.execute('SELECT name FROM students WHERE id = ?', (student_id,))
                name = cursor.fetchone()[0]
                conn.close()
                self.student_info_label.config(text=f"Найден студент: {name} (ID: {student_id})")
                self.current_reg_student_id = student_id
            else:
                self.student_info_label.config(text="Студент не найден")
                self.current_reg_student_id = None
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка поиска: {str(e)}")

    def load_meals_for_registration(self):
        """Load meals based on selected day"""
        day_choice = self.day_var.get()

        self.meals_listbox.delete(0, tk.END)

        conn = sqlite3.connect('skud.db')
        cursor = conn.cursor()

        if day_choice == "-1":
            cursor.execute('SELECT id, name, day_of_week FROM meals ORDER BY day_of_week, id')
        else:
            cursor.execute('SELECT id, name, day_of_week FROM meals WHERE day_of_week = ? ORDER BY id', (int(day_choice),))

        meals = cursor.fetchall()
        conn.close()

        day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

        for meal_id, name, day in meals:
            day_name = day_names[day] if day_choice == "-1" else ""
            display_text = f"{meal_id}: {name} {day_name}".strip()
            self.meals_listbox.insert(tk.END, display_text)
            # Store meal_id for later use
            if not hasattr(self, 'meal_ids'):
                self.meal_ids = []
            self.meal_ids.append(meal_id)

    def register_student(self):
        """Register student for selected meals"""
        if not hasattr(self, 'current_reg_student_id') or self.current_reg_student_id is None:
            messagebox.showwarning("Предупреждение", "Сначала найдите студента")
            return

        selected_indices = self.meals_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Предупреждение", "Выберите приемы пищи")
            return

        try:
            registered_count = 0
            for idx in selected_indices:
                if hasattr(self, 'meal_ids') and idx < len(self.meal_ids):
                    meal_id = self.meal_ids[idx]
                    add_registration(self.current_reg_student_id, meal_id)
                    registered_count += 1

            messagebox.showinfo("Успех", f"Студент зарегистрирован на {registered_count} приемов пищи")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось зарегистрировать: {str(e)}")

    def import_students_from_xlsx(self):
        """Import students from XLSX file"""
        filename = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            day_choice = self.day_var.get()
            if day_choice == "-1":
                d_list = list(range(7))  # Monday to Sunday
            else:
                d_list = [int(day_choice)]

            registered = 0
            skipped = 0
            for row in range(9, ws.max_row + 1):
                group = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=3).value  # ФИО in column 3
                if name and isinstance(name, str) and name.strip():
                    group_str = group.strip() if group and isinstance(group, str) else ""
                    student_id = find_student_by_name_group(name.strip(), group_str)
                    if student_id:
                        # Import registrations
                        for d in d_list:
                            for m in range(3):  # Breakfast, Lunch, Dinner
                                if day_choice == "-1":
                                    col = 4 + d * 3 + m
                                else:
                                    col = 4 + m
                                cell_value = ws.cell(row=row, column=col).value
                                if cell_value == 1 or str(cell_value).strip() in ['x', 'X', 'да', 'yes']:
                                    meal_id = d * 3 + m + 1
                                    add_registration(student_id, meal_id)
                                    registered += 1
                    else:
                        skipped += 1

            message = f"Зарегистрировано {registered} приемов пищи для существующих студентов"
            if skipped > 0:
                message += f"\nПропущено {skipped} студентов (не найдены в БД)"
            messagebox.showinfo("Успех", message)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать: {str(e)}")
