import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import openpyxl
from db.student import (
    finding_card, check_student, log_attendance, add_student,
    add_registration, get_all_students, update_student, delete_student,
    find_student_by_name_group
)

class StudentTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.create_ui()
        self.load_students()

    def create_ui(self):
        # Left panel - Student list
        left_frame = ttk.Frame(self)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        ttk.Label(left_frame, text="Список студентов:", font=('Arial', 12, 'bold')).pack(pady=5)

        # Treeview for students
        columns = ('ID', 'Имя', 'Карта', 'Группа')
        self.student_tree = ttk.Treeview(left_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.student_tree.heading(col, text=col)
            self.student_tree.column(col, width=100)
        self.student_tree.pack(fill=tk.BOTH, expand=True)

        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.student_tree.yview)
        self.student_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="Обновить список", command=self.load_students).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Добавить студента", command=self.add_student_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Импорт из Excel", command=self.import_students_info_from_xlsx).pack(side=tk.LEFT, padx=5)
        

        # Right panel - Actions
        right_frame = ttk.Frame(self)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        ttk.Label(right_frame, text="Действия:", font=('Arial', 12, 'bold')).pack(pady=5)

        action_frame = ttk.LabelFrame(right_frame, text="Выберите действие")
        action_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        ttk.Button(action_frame, text="Изменить студента", command=self.edit_student_dialog).pack(fill=tk.X, pady=5)
        ttk.Button(action_frame, text="Удалить студента", command=self.delete_student).pack(fill=tk.X, pady=5)

    def load_students(self):
        """Load all students into the treeview"""
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)

        students = get_all_students()
        for student in students:
            self.student_tree.insert('', tk.END, values=(student[0], student[1], student[2], student[3] if len(student) > 3 else ''))

    def add_student_dialog(self):
        """Dialog to add a new student"""
        name = simpledialog.askstring("Добавить студента", "Имя студента:")
        if not name:
            return

        group = simpledialog.askstring("Добавить студента", "Группа:")
        if not group:
            return

        card_id_str = simpledialog.askstring("Добавить студента", "ID карты:")
        if not card_id_str:
            return

        try:
            card_id = card_id_str
            add_student(name, card_id, group)
            messagebox.showinfo("Успех", f"Студент {name} из группы {group} добавлен с картой {card_id}")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось добавить студента: {str(e)}")

    def edit_student_dialog(self):
        """Dialog to edit selected student"""
        selected = self.student_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите студента для редактирования")
            return

        item = self.student_tree.item(selected[0])
        student_id, current_name, current_card, current_group = item['values']

        name = simpledialog.askstring("Изменить студента", "Новое имя:", initialvalue=current_name)
        group = simpledialog.askstring("Изменить студента", "Новая группа:", initialvalue=current_group)
        card_id_str = simpledialog.askstring("Изменить студента", "Новый ID карты:", initialvalue=str(current_card))

        try:
            card_id = card_id_str if card_id_str else None
            update_student(student_id, name, card_id, group)
            messagebox.showinfo("Успех", "Информация о студенте обновлена")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обновить студента: {str(e)}")

    def delete_student(self):
        """Delete selected student"""
        selected = self.student_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите студента для удаления")
            return

        if not messagebox.askyesno("Подтверждение", "Удалить студента и все связанные записи?"):
            return

        item = self.student_tree.item(selected[0])
        student_id = item['values'][0]

        try:
            delete_student(student_id)
            messagebox.showinfo("Успех", "Студент удален")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить студента: {str(e)}")

    def import_students_info_from_xlsx(self):
        """Import students from excel file"""
        filename = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            added = 0
            updated = 0
            for row in range(2, ws.max_row + 1):  # Skip header row
                uid = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value
                group = ws.cell(row=row, column=3).value

                if uid and name and group:
                    student_id = find_student_by_name_group(name.strip(), group.strip())
                    if student_id:
                        update_student(student_id, card_id=str(uid))
                        updated += 1
                    else:
                        add_student(name, str(uid), group)
                        added += 1

            messagebox.showinfo("Успех", f"Добавлено {added} студентов, обновлено {updated} студентов")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать: {str(e)}")
