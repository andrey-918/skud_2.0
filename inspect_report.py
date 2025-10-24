import openpyxl

wb = openpyxl.load_workbook('отчеты/report.xlsx')
ws = wb.active

print('Max rows:', ws.max_row, 'Max cols:', ws.max_column)

print('Headers:')
for col in range(1, ws.max_column + 1):
    print(f'Col {col}: {ws.cell(row=1, column=col).value}')

print('First 10 data rows:')
for row in range(2, min(12, ws.max_row + 1)):
    row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
    print(f'Row {row}: {row_data}')
