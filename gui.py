import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import sqlite3
from datetime import datetime
import openpyxl
from db.init_db import init_db
from db.student import (
    finding_card, check_student, log_attendance, add_student,
    add_registration, get_all_students, update_student, delete_student,
    find_student_by_name_group
)
from db.meals import get_current_meal, get_meal_name
from db.reports import get_all_attendance_records

class AttendanceSystemGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("СКУД")
        self.root.geometry("1000x700")

        # Initialize database
        init_db()

        # Create notebook (tabs)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Create tabs
        self.create_student_tab()
        self.create_registration_tab()
        self.create_reports_tab()
        self.create_attendance_tab()

    def create_student_tab(self):
        """Tab for managing students (add, edit, delete, view)"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Управление студентами")

        # Left panel - Student list
        left_frame = ttk.Frame(frame)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        ttk.Label(left_frame, text="Список студентов:", font=('Arial', 12, 'bold')).pack(pady=5)

        # Treeview for students
        columns = ('ID', 'Имя', 'Карта', 'Группа')
        self.student_tree = ttk.Treeview(left_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.student_tree.heading(col, text=col)
            self.student_tree.column(col, width=100)
        self.student_tree.pack(fill=tk.BOTH, expand=True)

        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(left_frame, orient=tk.VERTICAL, command=self.student_tree.yview)
        self.student_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=10)

        ttk.Button(btn_frame, text="Обновить список", command=self.load_students).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Добавить студента", command=self.add_student_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Импорт из Excel", command=self.import_students_info_from_xlsx).pack(side=tk.LEFT, padx=5)
        

        # Right panel - Actions
        right_frame = ttk.Frame(frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        ttk.Label(right_frame, text="Действия:", font=('Arial', 12, 'bold')).pack(pady=5)

        action_frame = ttk.LabelFrame(right_frame, text="Выберите действие")
        action_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        ttk.Button(action_frame, text="Изменить студента", command=self.edit_student_dialog).pack(fill=tk.X, pady=5)
        ttk.Button(action_frame, text="Удалить студента", command=self.delete_student).pack(fill=tk.X, pady=5)

        # Load initial data
        self.load_students()

    def create_registration_tab(self):
        """Tab for registering students for meals"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Регистрация на приемы пищи")

        # Student selection
        student_frame = ttk.LabelFrame(frame, text="Выбор студента")
        student_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(student_frame, text="ID карты студента:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.reg_card_id = ttk.Entry(student_frame)
        self.reg_card_id.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)

        ttk.Button(student_frame, text="Найти студента", command=self.find_student_for_registration).grid(row=0, column=2, padx=5, pady=5)

        self.student_info_label = ttk.Label(student_frame, text="")
        self.student_info_label.grid(row=1, column=0, columnspan=3, padx=5, pady=5, sticky=tk.W)

        # Day selection
        day_frame = ttk.LabelFrame(frame, text="Выбор дня")
        day_frame.pack(fill=tk.X, padx=10, pady=10)

        self.day_var = tk.StringVar(value="-1")
        days = [("-1", "Все дни"), ("0", "Понедельник"), ("1", "Вторник"), ("2", "Среда"),
                ("3", "Четверг"), ("4", "Пятница"), ("5", "Суббота"), ("6", "Воскресенье")]

        for i, (val, text) in enumerate(days):
            ttk.Radiobutton(day_frame, text=text, variable=self.day_var, value=val).grid(row=i//4, column=i%4, padx=5, pady=5, sticky=tk.W)

        # Meals selection
        meals_frame = ttk.LabelFrame(frame, text="Выбор приемов пищи")
        meals_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.meals_listbox = tk.Listbox(meals_frame, selectmode=tk.MULTIPLE, height=10)
        self.meals_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(meals_frame, orient=tk.VERTICAL, command=self.meals_listbox.yview)
        self.meals_listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(btn_frame, text="Загрузить приемы пищи", command=self.load_meals_for_registration).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Зарегистрировать", command=self.register_student).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Импорт из XLSX", command=self.import_students_from_xlsx).pack(side=tk.LEFT, padx=5)

    def create_reports_tab(self):
        """Tab for exporting attendance reports to Excel"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Отчеты")

        # Threshold for percentage coloring
        threshold_frame = ttk.LabelFrame(frame, text="Настройки")
        threshold_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(threshold_frame, text="Порог процента для окраски (%):").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.percentage_threshold = ttk.Entry(threshold_frame)
        self.percentage_threshold.insert(0, "50")
        self.percentage_threshold.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)

        # Export button for full report
        ttk.Button(frame, text="Экспорт общей таблицы в Excel", command=self.export_all_report_to_excel).pack(pady=20)

    def create_attendance_tab(self):
        """Tab for checking attendance (main functionality)"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Проверка посещаемости")

        # Current meal info
        info_frame = ttk.LabelFrame(frame, text="Текущий прием пищи")
        info_frame.pack(fill=tk.X, padx=10, pady=10)

        self.current_meal_label = ttk.Label(info_frame, text="Определение текущего приема пищи...")
        self.current_meal_label.pack(padx=10, pady=10)

        ttk.Button(info_frame, text="Обновить", command=self.update_current_meal).pack(pady=5)

        # Attendance check
        check_frame = ttk.LabelFrame(frame, text="Проверка доступа")
        check_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        ttk.Label(check_frame, text="ID карты студента:", font=('Arial', 12)).pack(pady=10)

        self.attendance_card_id = ttk.Entry(check_frame, font=('Arial', 14))
        self.attendance_card_id.pack(pady=5, ipady=5)

        ttk.Button(check_frame, text="Проверить доступ", command=self.check_attendance).pack(pady=20)

        self.attendance_result = ttk.Label(check_frame, text="", font=('Arial', 16, 'bold'))
        self.attendance_result.pack(pady=20)

        # Initialize current meal
        self.update_current_meal()

    # Student management methods
    def load_students(self):
        """Load all students into the treeview"""
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)

        students = get_all_students()
        for student in students:
            self.student_tree.insert('', tk.END, values=(student[0], student[1], student[2], student[3] if len(student) > 3 else ''))

    def add_student_dialog(self):
        """Dialog to add a new student"""
        name = simpledialog.askstring("Добавить студента", "Имя студента:")
        if not name:
            return

        group = simpledialog.askstring("Добавить студента", "Группа:")
        if not group:
            return

        card_id_str = simpledialog.askstring("Добавить студента", "ID карты:")
        if not card_id_str:
            return

        try:
            card_id = card_id_str
            add_student(name, card_id, group)
            messagebox.showinfo("Успех", f"Студент {name} из группы {group} добавлен с картой {card_id}")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось добавить студента: {str(e)}")

    def edit_student_dialog(self):
        """Dialog to edit selected student"""
        selected = self.student_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите студента для редактирования")
            return

        item = self.student_tree.item(selected[0])
        student_id, current_name, current_card, current_group = item['values']

        name = simpledialog.askstring("Изменить студента", "Новое имя:", initialvalue=current_name)
        group = simpledialog.askstring("Изменить студента", "Новая группа:", initialvalue=current_group)
        card_id_str = simpledialog.askstring("Изменить студента", "Новый ID карты:", initialvalue=str(current_card))

        try:
            card_id = card_id_str if card_id_str else None
            update_student(student_id, name, card_id, group)
            messagebox.showinfo("Успех", "Информация о студенте обновлена")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обновить студента: {str(e)}")

    def delete_student(self):
        """Delete selected student"""
        selected = self.student_tree.selection()
        if not selected:
            messagebox.showwarning("Предупреждение", "Выберите студента для удаления")
            return

        if not messagebox.askyesno("Подтверждение", "Удалить студента и все связанные записи?"):
            return

        item = self.student_tree.item(selected[0])
        student_id = item['values'][0]

        try:
            delete_student(student_id)
            messagebox.showinfo("Успех", "Студент удален")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить студента: {str(e)}")

    def import_students_from_xlsx(self):
        """Import students from XLSX file"""
        filename = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            day_choice = self.day_var.get()
            if day_choice == "-1":
                d_list = list(range(7))  # Monday to Sunday
            else:
                d_list = [int(day_choice)]

            registered = 0
            skipped = 0
            for row in range(9, ws.max_row + 1):
                group = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=3).value  # ФИО in column 3
                if name and isinstance(name, str) and name.strip():
                    group_str = group.strip() if group and isinstance(group, str) else ""
                    student_id = find_student_by_name_group(name.strip(), group_str)
                    if student_id:
                        # Import registrations
                        for d in d_list:
                            for m in range(3):  # Breakfast, Lunch, Dinner
                                if day_choice == "-1":
                                    col = 4 + d * 3 + m
                                else:
                                    col = 4 + m
                                cell_value = ws.cell(row=row, column=col).value
                                if cell_value == 1 or str(cell_value).strip() in ['x', 'X', 'да', 'yes']:
                                    meal_id = d * 3 + m + 1
                                    add_registration(student_id, meal_id)
                                    registered += 1
                    else:
                        skipped += 1

            message = f"Зарегистрировано {registered} приемов пищи для существующих студентов"
            if skipped > 0:
                message += f"\nПропущено {skipped} студентов (не найдены в БД)"
            messagebox.showinfo("Успех", message)
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать: {str(e)}")

    def import_students_info_from_xlsx(self):
        """Import students from excel file"""
        filename = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            added = 0
            updated = 0
            for row in range(2, ws.max_row + 1):  # Skip header row
                uid = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value
                group = ws.cell(row=row, column=3).value

                if uid and name and group:
                    student_id = find_student_by_name_group(name.strip(), group.strip())
                    if student_id:
                        update_student(student_id, card_id=str(uid))
                        updated += 1
                    else:
                        add_student(name, str(uid), group)
                        added += 1

            messagebox.showinfo("Успех", f"Добавлено {added} студентов, обновлено {updated} студентов")
            self.load_students()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать: {str(e)}")

    # Registration methods
    def find_student_for_registration(self):
        """Find student by card ID for registration"""
        card_id_str = self.reg_card_id.get()
        if not card_id_str:
            messagebox.showwarning("Предупреждение", "Введите ID карты")
            return

        try:
            card_id = card_id_str
            student_id = finding_card(card_id)
            if student_id:
                # Get student name
                conn = sqlite3.connect('skud.db')
                cursor = conn.cursor()
                cursor.execute('SELECT name FROM students WHERE id = ?', (student_id,))
                name = cursor.fetchone()[0]
                conn.close()
                self.student_info_label.config(text=f"Найден студент: {name} (ID: {student_id})")
                self.current_reg_student_id = student_id
            else:
                self.student_info_label.config(text="Студент не найден")
                self.current_reg_student_id = None
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка поиска: {str(e)}")

    def load_meals_for_registration(self):
        """Load meals based on selected day"""
        day_choice = self.day_var.get()

        self.meals_listbox.delete(0, tk.END)

        conn = sqlite3.connect('skud.db')
        cursor = conn.cursor()

        if day_choice == "-1":
            cursor.execute('SELECT id, name, day_of_week FROM meals ORDER BY day_of_week, id')
        else:
            cursor.execute('SELECT id, name, day_of_week FROM meals WHERE day_of_week = ? ORDER BY id', (int(day_choice),))

        meals = cursor.fetchall()
        conn.close()

        day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

        for meal_id, name, day in meals:
            day_name = day_names[day] if day_choice == "-1" else ""
            display_text = f"{meal_id}: {name} {day_name}".strip()
            self.meals_listbox.insert(tk.END, display_text)
            # Store meal_id for later use
            if not hasattr(self, 'meal_ids'):
                self.meal_ids = []
            self.meal_ids.append(meal_id)

    def register_student(self):
        """Register student for selected meals"""
        if not hasattr(self, 'current_reg_student_id') or self.current_reg_student_id is None:
            messagebox.showwarning("Предупреждение", "Сначала найдите студента")
            return

        selected_indices = self.meals_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Предупреждение", "Выберите приемы пищи")
            return

        try:
            registered_count = 0
            for idx in selected_indices:
                if hasattr(self, 'meal_ids') and idx < len(self.meal_ids):
                    meal_id = self.meal_ids[idx]
                    add_registration(self.current_reg_student_id, meal_id)
                    registered_count += 1

            messagebox.showinfo("Успех", f"Студент зарегистрирован на {registered_count} приемов пищи")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось зарегистрировать: {str(e)}")

    # Reports methods
    def export_all_report_to_excel(self):
        """Export comprehensive attendance report to Excel in wide format"""
        try:
            # Get all attendance records
            records = get_all_attendance_records()
            if not records:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
                return

            # Group records by student
            from collections import defaultdict
            student_data = defaultdict(lambda: defaultdict(dict))
            student_info = {}

            for record in records:
                student_name = record['student_name']
                meal_id = record['meal_id']
                day_of_week = record['day_of_week']
                status = record['status']

                # Calculate meal type: 0=Breakfast, 1=Lunch, 2=Dinner
                meal_type = (meal_id - 1) % 3
                student_data[student_name][day_of_week][meal_type] = status

                # Store student info (assuming all records have same info)
                if student_name not in student_info:
                    # Get student details
                    conn = sqlite3.connect('skud.db')
                    cursor = conn.cursor()
                    cursor.execute('SELECT id, card_id, group_name FROM students WHERE name = ?', (student_name,))
                    result = cursor.fetchone()
                    conn.close()
                    if result:
                        student_info[student_name] = {'id': result[0], 'card_id': result[1], 'group': result[2]}

            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Общая таблица посещаемости"

            # Headers similar to report.xlsx
            ws['A1'] = "График обеспечением питания студентов Морского института"
            ws['A5'] = "Дата"
            ws['C5'] = f"{datetime.now().strftime('%d.%m.%Y')}"
            ws['A6'] = "День недели"
            ws['D6'] = "Понедельник"
            ws['G6'] = "Вторник"
            ws['J6'] = "Среда"
            ws['M6'] = "Четверг"
            ws['P6'] = "Пятница"
            ws['S6'] = "Суббота"
            ws['V6'] = "Воскресенье"
            ws['A7'] = "Локация"
            ws['D7'] = "Гоголя"
            ws['G7'] = "Гоголя"
            ws['J7'] = "Гоголя"
            ws['M7'] = "Гоголя"
            ws['P7'] = "Гоголя"
            ws['S7'] = "Гоголя"
            ws['V7'] = "Гоголя"
            ws['A8'] = "Учебная группа"
            ws['B8'] = "№п/п"
            ws['C8'] = "ФИО"
            # Meal headers: З (Breakfast), О (Lunch), У (Dinner)
            col = 4
            for day in range(7):
                ws.cell(row=8, column=col).value = "З"
                ws.cell(row=8, column=col+1).value = "О"
                ws.cell(row=8, column=col+2).value = "У"
                col += 3

            # Additional columns
            ws.cell(row=8, column=22).value = "Всего заявок"
            ws.cell(row=8, column=23).value = "Питание по заявке"
            ws.cell(row=8, column=24).value = "Питание без заявки"
            ws.cell(row=8, column=25).value = "Процент посещений"

            # Data rows
            row = 9
            for idx, student_name in enumerate(sorted(student_data.keys()), start=1):
                info = student_info.get(student_name, {'group': '', 'id': '', 'card_id': ''})
                ws.cell(row=row, column=1).value = info['group'] or ""
                ws.cell(row=row, column=2).value = idx
                ws.cell(row=row, column=3).value = student_name

                total_registered = 0
                came_with_registration = 0
                came_without_registration = 0

                col = 4
                for day in range(7):
                    for meal_type in range(3):  # 0: Breakfast, 1: Lunch, 2: Dinner
                        status = student_data[student_name][day].get(meal_type, 'not_registered')
                        cell = ws.cell(row=row, column=col)
                        if status == 'came':
                            cell.value = 1
                            came_with_registration += 1
                            total_registered += 1
                        elif status == 'didnt_come':
                            cell.value = 1
                            cell.fill = yellow_fill
                            total_registered += 1
                        elif status == 'came_without_registration':
                            cell.value = 1
                            came_without_registration += 1
                        elif status == 'not_registered':
                            cell.value = 0
                        col += 1

                # Calculate total registered (didnt_come + came_with_registration)
                total_registered += came_with_registration

                # Additional columns
                cell_22 = ws.cell(row=row, column=22)
                cell_22.value = total_registered

                cell_23 = ws.cell(row=row, column=23)
                cell_23.value = came_with_registration

                cell_24 = ws.cell(row=row, column=24)
                cell_24.value = came_without_registration

                cell_25 = ws.cell(row=row, column=25)
                if total_registered > 0:
                    percentage = (came_with_registration / total_registered) * 100
                    cell_25.value = round(percentage, 2)
                    try:
                        threshold = float(self.percentage_threshold.get())
                        if percentage < threshold:
                            cell_25.fill = red_fill
                    except ValueError:
                        pass  # Ignore invalid threshold
                else:
                    cell_25.value = 0

                row += 1

            filename = f"отчеты/Общая_таблица_посещаемости_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            wb.save(filename)

            messagebox.showinfo("Успех", f"Отчет экспортирован в {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать отчет: {str(e)}")

    # Attendance methods
    def update_current_meal(self):
        """Update current meal information"""
        current_meal = get_current_meal()
        if current_meal is None:
            self.current_meal_label.config(text="Сейчас нет времени приема пищи")
            self.current_meal_id = None
        else:
            meal_name = get_meal_name(current_meal)
            day_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
            day_name = day_names[datetime.now().weekday()]
            self.current_meal_label.config(text=f"Текущий прием: {meal_name} ({day_name})")
            self.current_meal_id = current_meal

    def check_attendance(self):
        """Check student attendance for current meal"""
        if self.current_meal_id is None:
            self.attendance_result.config(text="Сейчас нет времени приема пищи", foreground="red")
            return

        card_id_str = self.attendance_card_id.get()
        if not card_id_str:
            self.attendance_result.config(text="Введите ID карты", foreground="red")
            return

        try:
            card_id = card_id_str
            status = check_student(card_id, self.current_meal_id)
            if status == 'registered':
                student_id = finding_card(card_id)
                log_attendance(student_id, self.current_meal_id, 'came')
                self.attendance_result.config(text="ДОСТУП РАЗРЕШЕН", foreground="green")
            elif status == 'not_registered':
                student_id = finding_card(card_id)
                log_attendance(student_id, self.current_meal_id, 'came_without_registration')
                self.attendance_result.config(text="ДОСТУП РАЗРЕШЕН (БЕЗ ЗАПИСИ)", foreground="orange")
            else:  # unknown_card
                self.attendance_result.config(text="ДОСТУП ЗАПРЕЩЕН", foreground="red")
        except Exception as e:
            self.attendance_result.config(text=f"Ошибка: {str(e)}", foreground="red")

        # Clear the input
        self.attendance_card_id.delete(0, tk.END)


def main():
    root = tk.Tk()
    app = AttendanceSystemGUI(root)
    root.mainloop()
